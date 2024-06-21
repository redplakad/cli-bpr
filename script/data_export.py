import os
import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from datetime import datetime
import random
import string
from colorama import init, Fore, Style

# Function to generate random string of given length
def random_string(length):
    letters = string.ascii_lowercase
    return ''.join(random.choice(letters) for i in range(length))

# Function to export specific columns from 'loan' table to Excel
def export_nominatif(conn):
    try:
        # Specify columns to fetch from 'loan' table
        columns_to_fetch = [
            'NOMOR_REKENING', 'NO_CIF', 'TITIPAN_EFEKTIF', 'NAMA_NASABAH', 'POKOK_PINJAMAN', 'TUNGGAKAN_POKOK', 
            'TUNGGAKAN_BUNGA', 'ANGSURAN_TOTAL', 'KODE_KOLEK', 'TGL_AKHIR_BAYAR', 'JML_HARI_TUNGGAKAN', 
            'AO', 'TEMPAT_BEKERJA', 'ALAMAT', 'KELURAHAN', 'KECAMATAN', 'NO_HP', 'BGA', 'CADANGAN_PPAP'
        ]

        cursor = conn.cursor()
        select_query = f"SELECT {', '.join(columns_to_fetch)} FROM loan"
        cursor.execute(select_query)

        # Fetch all rows from the cursor
        rows = cursor.fetchall()

        # Convert to DataFrame
        df = pd.DataFrame(rows, columns=columns_to_fetch)

        # Add 'No' column for row numbering
        df.insert(0, 'No', range(1, len(df) + 1))

        # Modify 'NOMOR_REKENING' and 'NO_CIF' to include a leading apostrophe
        df['NOMOR_REKENING'] = df['NOMOR_REKENING'].apply(lambda x: f"00{x}")
        df['NO_CIF'] = df['NO_CIF'].apply(lambda x: f"00{x}")

        # Create Excel workbook and sheet
        wb = Workbook()
        ws = wb.active
        ws.title = 'Nominatif'

        # Define header style (black background, white text)
        header_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')

        # Write header row with style
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Write data rows
        for row_idx, row_data in enumerate(df.values, start=2):
            for col_idx, cell_value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                # Apply text format for 'NOMOR_REKENING' and 'NO_CIF' columns
                if df.columns[col_idx - 1] in ['NOMOR_REKENING', 'NO_CIF']:
                    cell.number_format = '@'
                # Apply comma-separated format for specific columns
                if df.columns[col_idx - 1] in ['POKOK_PINJAMAN', 'TUNGGAKAN_POKOK', 'TUNGGAKAN_BUNGA', 'ANGSURAN_TOTAL']:
                    cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1

        # Adjust column widths to fit content
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # Generate unique timestamp string
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Construct file name with timestamp
        excel_file = f"nominatif_{timestamp}.xlsx"

        # Save workbook to Excel file
        wb.save(excel_file)

        # Clear the terminal
        os.system('cls' if os.name == 'nt' else 'clear')

        print(Fore.GREEN + "=" * 95)
        print(Fore.GREEN + f"  ++++ Data Nominatif berhasil diekspor ke {excel_file}. ++++  ")
        print(Fore.GREEN + "=" * 95)

    except sqlite3.Error as e:
        print(f"Error exporting data: {e}")
    finally:
        cursor.close()

# Example usage
if __name__ == "__main__":
    # Initialize colorama
    init(autoreset=True)

    # SQLite database file
    db_file = 'database.db'
    
    # Create connection to SQLite database
    conn = sqlite3.connect(db_file)
    
    # Call the export function
    export_nominatif(conn)
    
    # Close database connection
    conn.close()
